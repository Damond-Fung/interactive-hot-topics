#!/usr/bin/env python3
"""Generate interactive forum hot topic report for the TRAE forum."""

from __future__ import annotations

import argparse
import json
import re
import time as time_module
from collections import Counter
from datetime import datetime, time, timedelta, timezone
from html import unescape
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


DEFAULT_FORUM_BASE_URL = "https://forum.trae.cn"
DEFAULT_TIMEZONE = "Asia/Shanghai"
DEFAULT_OUTPUT_DIR = "exports"
USER_AGENT = "interactive-hot-topics-report/1.0"
INTERACTIVE_CATEGORY_ID = 11
INTERACTIVE_CATEGORY_NAME = "互动交流"
CHINESE_STOPWORDS = {
    "这个",
    "那个",
    "一下",
    "问题",
    "请问",
    "求助",
    "大家",
    "感觉",
    "如何",
    "怎么",
    "今天",
    "本周",
    "论坛",
    "社区",
    "帖子",
    "话题",
    "交流",
    "互动",
    "功能",
    "设置",
    "使用",
    "可以",
    "就是",
    "一个",
    "什么",
}
ASCII_STOPWORDS = {
    "trae",
    "the",
    "and",
    "for",
    "with",
    "from",
    "that",
    "this",
    "what",
    "how",
    "when",
    "why",
    "you",
    "use",
    "using",
    "issue",
    "problem",
}
TOPIC_PATTERNS = {
    "排队": ["排队", "等待", "繁忙", "busy", "队列", "稍后重试"],
    "Skill": ["skill", "skills"],
    "SOLO": ["solo"],
    "MCP": ["mcp"],
    "国际版/额度": ["国际版", "pro", "bonus", "额度", "用量", "token", "缩水"],
    "模型支持": ["glm-5.1", "glm 5.1", "kimi coding plan", "agent team", "seed-2.0", "seed 2.0"],
    "自动继续": ["自动继续", "继续", "continue"],
    "远程服务器": ["远程服务器", "ssh", "服务器"],
    "收费/发票": ["收费", "计费", "发票", "优速通"],
}
NEGATIVE_PATTERNS = {
    "TRAE CN排队": ["排队", "等待", "繁忙", "busy", "稍后重试"],
    "TRAE 国际版实际用量缩水": ["国际版", "pro", "bonus", "额度", "用量", "token", "缩水"],
    "中断/卡住/无输出": ["中断", "卡住", "无输出", "没有输出", "没结果", "提前退出", "失败", "报错", "异常"],
    "收费/计费/发票": ["收费", "计费", "发票", "优速通"],
}
SUGGESTION_PATTERNS = {
    "支持 GLM-5.1": ["glm-5.1", "glm 5.1", "支持glm", "支持 glm"],
    "支持 Kimi Coding Plan": ["kimi coding plan", "kimi", "coding plan"],
    "Agent Team": ["agent team", "团队协作", "多人协作"],
    "自动继续": ["自动继续", "继续", "continue"],
}
SUGGESTION_PREFIX_RE = re.compile(r"^(支持|建议|加入|增加|新增|接入|适配)\s*([A-Za-z0-9.+#\-/ ]{2,24}|[\u4e00-\u9fffA-Za-z0-9.+#\-/ ]{2,18})")
HTML_TAG_RE = re.compile(r"<[^>]+>")
HTML_CODE_RE = re.compile(r"<pre[\s\S]*?</pre>|<code[\s\S]*?</code>", re.IGNORECASE)
WHITESPACE_RE = re.compile(r"\s+")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate interactive hot topics Excel report.")
    parser.add_argument("--forum-base-url", default=DEFAULT_FORUM_BASE_URL)
    parser.add_argument("--timezone", default=DEFAULT_TIMEZONE)
    parser.add_argument(
        "--time-preset",
        choices=["last-week", "this-week", "last-7-days"],
        default="last-week",
    )
    parser.add_argument("--start-date", help="YYYY-MM-DD, inclusive")
    parser.add_argument("--end-date", help="YYYY-MM-DD, inclusive")
    parser.add_argument("--top-n", type=int, default=5)
    parser.add_argument("--output", help="Target .xlsx path")
    parser.add_argument("--max-pages", type=int, default=30)
    parser.add_argument("--ai-results", help="Optional JSON file that contains AI analysis results keyed by topic_id.")
    parser.add_argument("--export-ai-template", action="store_true", help="Export an AI results template JSON next to the report.")
    return parser.parse_args()


def parse_local_date(value: str) -> datetime:
    return datetime.strptime(value, "%Y-%m-%d")


def compute_window(args: argparse.Namespace) -> tuple[datetime, datetime, str]:
    tz = ZoneInfo(args.timezone)
    now = datetime.now(tz)

    if args.start_date or args.end_date:
        if not (args.start_date and args.end_date):
            raise ValueError("Custom time range requires both --start-date and --end-date.")
        start_local = datetime.combine(parse_local_date(args.start_date).date(), time.min, tz)
        end_local = datetime.combine(parse_local_date(args.end_date).date(), time.max, tz)
        return start_local, end_local, f"{args.start_date} to {args.end_date}"

    if args.time_preset == "last-week":
        this_week_start = datetime.combine((now - timedelta(days=now.weekday())).date(), time.min, tz)
        return this_week_start - timedelta(days=7), this_week_start - timedelta(microseconds=1), "last-week"

    if args.time_preset == "this-week":
        start_local = datetime.combine((now - timedelta(days=now.weekday())).date(), time.min, tz)
        return start_local, now, "this-week"

    if args.time_preset == "last-7-days":
        end_local = now
        start_local = datetime.combine((now - timedelta(days=6)).date(), time.min, tz)
        return start_local, end_local, "last-7-days"

    raise ValueError(f"Unsupported time preset: {args.time_preset}")


def parse_iso_datetime(value: str) -> datetime:
    return datetime.fromisoformat(value.replace("Z", "+00:00"))


def dedupe_preserve_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for item in items:
        if not item or item in seen:
            continue
        seen.add(item)
        output.append(item)
    return output


def strip_html(value: str) -> str:
    no_code = HTML_CODE_RE.sub(" ", value or "")
    plain = HTML_TAG_RE.sub(" ", no_code)
    return WHITESPACE_RE.sub(" ", unescape(plain)).strip()


def normalize_analysis_text(value: str) -> str:
    lowered = strip_html(value).lower()
    lowered = re.sub(r"[\[\]【】（）()\"'“”‘’`~!@#$%^&*+=|\\/:;,.?，。！？：；、<>-]+", " ", lowered)
    return WHITESPACE_RE.sub(" ", lowered).strip()


def compact_analysis_text(value: str) -> str:
    return normalize_analysis_text(value).replace(" ", "")


def contains_term(normalized_text: str, compact_text: str, term: str) -> bool:
    normalized_term = normalize_analysis_text(term)
    if not normalized_term:
        return False

    if re.search(r"[\u4e00-\u9fff]", term):
        return normalized_term.replace(" ", "") in compact_text

    return re.search(rf"(?<![a-z0-9]){re.escape(normalized_term)}(?![a-z0-9])", normalized_text) is not None


def title_fragments(text: str) -> list[str]:
    fragments: list[str] = []
    for part in re.split(r"[^\w\u4e00-\u9fff.+#/-]+", strip_html(text)):
        token = part.strip().strip("-_/#")
        if len(token) < 2:
            continue
        lowered = token.lower()
        if token.isascii():
            if len(lowered) < 3 or lowered in ASCII_STOPWORDS:
                continue
            fragments.append(token.upper() if lowered == "mcp" else token)
            continue
        if len(token) > 12 or token in CHINESE_STOPWORDS:
            continue
        fragments.append(token)
    return dedupe_preserve_order(fragments)


def extract_pattern_terms(title: str, texts: list[str]) -> list[str]:
    combined = " ".join([title, *texts])
    normalized = normalize_analysis_text(combined)
    compact = compact_analysis_text(combined)
    matched: list[str] = []

    for canonical, variants in TOPIC_PATTERNS.items():
        if any(contains_term(normalized, compact, variant) for variant in variants):
            matched.append(canonical)

    return dedupe_preserve_order(matched)


def extract_keywords(title: str, texts: list[str]) -> list[str]:
    keywords: list[str] = []
    keywords.extend(extract_pattern_terms(title, texts))
    keywords.extend(title_fragments(title))
    for text in texts[:2]:
        keywords.extend(title_fragments(text))
    return dedupe_preserve_order(keywords)[:8]


def extract_negative_signals(text: str) -> list[str]:
    normalized = normalize_analysis_text(text)
    compact = normalized.replace(" ", "")
    signals: list[str] = []

    if any(contains_term(normalized, compact, variant) for variant in NEGATIVE_PATTERNS["TRAE CN排队"]):
        signals.append("TRAE CN排队")

    intl_hit = any(contains_term(normalized, compact, variant) for variant in ["国际版", "pro", "bonus"])
    quota_hit = any(contains_term(normalized, compact, variant) for variant in ["额度", "用量", "token", "缩水"])
    if intl_hit and quota_hit:
        signals.append("TRAE 国际版实际用量缩水")

    if any(contains_term(normalized, compact, variant) for variant in NEGATIVE_PATTERNS["中断/卡住/无输出"]):
        signals.append("中断/卡住/无输出")

    if any(contains_term(normalized, compact, variant) for variant in NEGATIVE_PATTERNS["收费/计费/发票"]):
        signals.append("收费/计费/发票")

    return dedupe_preserve_order(signals)


def cleanup_suggestion_label(value: str) -> str:
    cleaned = WHITESPACE_RE.sub(" ", value.replace("“", " ").replace("”", " ").replace('"', " ")).strip(" -_/")
    if not cleaned:
        return ""
    if cleaned.isascii():
        parts = cleaned.split()
        normalized_parts = [part.upper() if part.lower() in {"glm", "mcp"} else part for part in parts]
        return " ".join(normalized_parts)
    return cleaned


def extract_product_suggestions(text: str) -> list[str]:
    title_text = strip_html(text)
    normalized = normalize_analysis_text(title_text)
    compact = normalized.replace(" ", "")
    suggestions: list[str] = []

    for canonical, variants in SUGGESTION_PATTERNS.items():
        if any(contains_term(normalized, compact, variant) for variant in variants):
            suggestions.append(canonical)

    for prefix, body in SUGGESTION_PREFIX_RE.findall(title_text):
        label = cleanup_suggestion_label(body)
        if not label:
            continue
        if prefix == "支持":
            suggestions.append(f"支持 {label}")
        else:
            suggestions.append(label)

    return dedupe_preserve_order(suggestions)


def clip_text(value: str, limit: int) -> str:
    text = strip_html(value)
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 3)].rstrip() + "..."


def build_post_blocks(posts: list[dict[str, Any]], timezone_name: str) -> tuple[list[dict[str, Any]], str]:
    tz = ZoneInfo(timezone_name)
    blocks: list[dict[str, Any]] = []
    lines: list[str] = []

    for post in posts:
        created_local = parse_iso_datetime(post["created_at"]).astimezone(tz).strftime("%Y-%m-%d %H:%M:%S")
        author = post.get("username") or "-"
        post_number = int(post.get("post_number") or 0)
        body = strip_html(post.get("cooked") or "")
        block = {
            "post_number": post_number,
            "author": author,
            "created_at": created_local,
            "content": body,
        }
        blocks.append(block)
        lines.append(f"[#{post_number}] {author} {created_local}\n{body}")

    return blocks, "\n\n".join(lines)


def build_agent_analysis_input(
    *,
    topic_id: int,
    title: str,
    topic_url: str,
    created_at: str,
    views: int,
    reply_count: int,
    weekly_post_count: int,
    weekly_reply_count: int,
    thread_text: str,
) -> str:
    lines = [
        f"topic_id: {topic_id}",
        f"标题: {title}",
        f"链接: {topic_url}",
        f"创建时间: {created_at}",
        f"浏览量: {views}",
        f"总回复数: {reply_count}",
        f"窗口内帖子数: {weekly_post_count}",
        f"窗口内回帖数: {weekly_reply_count}",
        "",
        "请基于标题和全帖内容判断：",
        "1. 话题分类",
        "2. 是否为负向反馈（是/否）",
        "3. 是否为产品建议（是/否）",
        "4. 话题核心内容（1句话）",
        "",
        "全帖内容：",
        thread_text,
    ]
    return "\n".join(lines)


def request_json(url: str, *, retries: int = 3, timeout: int = 30) -> dict[str, Any]:
    last_error: Exception | None = None
    for attempt in range(retries):
        req = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "application/json"})
        try:
            with urlopen(req, timeout=timeout) as resp:
                return json.load(resp)
        except HTTPError as exc:
            if exc.code in {404, 410}:
                raise
            last_error = exc
        except URLError as exc:
            last_error = exc

        if attempt < retries - 1:
            time_module.sleep(min(2 * (attempt + 1), 5))

    if last_error:
        raise last_error
    raise RuntimeError(f"Failed to fetch JSON from {url}")


def build_listing_url(source_url: str, page: int, order: str) -> str:
    parsed = urlparse(source_url)
    path = parsed.path.rstrip("/")
    if path.endswith("/l/latest"):
        listing_path = f"{path}.json"
    elif path.endswith("/l/latest.json"):
        listing_path = path
    elif path.endswith(".json") and "/l/" not in path:
        listing_path = f"{path[:-5]}/l/latest.json"
    else:
        listing_path = f"{path}/l/latest.json"

    query_items = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query_items["order"] = order
    if page > 0:
        query_items["page"] = str(page)
    else:
        query_items.pop("page", None)

    return urlunparse(
        (
            parsed.scheme,
            parsed.netloc,
            listing_path,
            parsed.params,
            urlencode(query_items),
            parsed.fragment,
        )
    )


def build_category_url(base_url: str, category_id: int) -> str:
    return f"{base_url.rstrip('/')}/c/{category_id}"


def fetch_topic_posts(topic_id: int, base_url: str) -> dict[str, Any]:
    topic = request_json(f"{base_url.rstrip('/')}/t/-/{topic_id}.json")
    posts = list(topic["post_stream"]["posts"])
    existing_ids = {post["id"] for post in posts}
    remaining = [post_id for post_id in topic["post_stream"]["stream"] if post_id not in existing_ids]

    for offset in range(0, len(remaining), 20):
        chunk_ids = remaining[offset : offset + 20]
        query = "&".join(f"post_ids%5B%5D={post_id}" for post_id in chunk_ids)
        chunk = request_json(f"{base_url.rstrip('/')}/t/{topic_id}/posts.json?{query}")
        posts.extend(chunk.get("post_stream", {}).get("posts", chunk.get("posts", [])))

    topic["all_posts"] = posts
    return topic


def collect_created_topics(base_url: str, category_id: int, start_utc: datetime, end_utc: datetime, max_pages: int) -> list[dict[str, Any]]:
    source_url = build_category_url(base_url, category_id)
    collected: list[dict[str, Any]] = []
    seen_ids: set[int] = set()

    for page in range(max_pages):
        payload = request_json(build_listing_url(source_url, page, "created"))
        batch = payload.get("topic_list", {}).get("topics", [])
        if not batch:
            break

        oldest_dt: datetime | None = None
        for topic in batch:
            topic_id = topic["id"]
            if topic_id in seen_ids:
                continue
            seen_ids.add(topic_id)
            created_at = parse_iso_datetime(topic["created_at"])
            topic["created_at_dt"] = created_at
            collected.append(topic)

            if oldest_dt is None or created_at < oldest_dt:
                oldest_dt = created_at

        if oldest_dt and oldest_dt < start_utc:
            break

    return [topic for topic in collected if start_utc <= topic["created_at_dt"] <= end_utc]


def topic_posts_within_window(topic: dict[str, Any], start_utc: datetime, end_utc: datetime) -> list[dict[str, Any]]:
    posts: list[dict[str, Any]] = []
    for post in topic.get("all_posts", []):
        created_at = parse_iso_datetime(post["created_at"])
        if not (start_utc <= created_at <= end_utc):
            continue
        if post.get("post_type", 1) != 1:
            continue
        if post.get("hidden") or post.get("user_deleted"):
            continue
        posts.append({**post, "created_at_dt": created_at})
    return posts


def update_bucket(
    buckets: dict[str, dict[str, Any]],
    label: str,
    detail_row: dict[str, Any],
    *,
    score: int,
    mention_weight: int,
) -> None:
    bucket = buckets.setdefault(
        label,
        {
            "label": label,
            "topic_ids": set(),
            "topic_count": 0,
            "reply_total": 0,
            "vote_total": 0,
            "mention_count": 0,
            "score": 0,
            "sample_titles": [],
        },
    )
    bucket["mention_count"] += mention_weight
    bucket["score"] += score
    topic_id = detail_row["topic_id"]
    if topic_id in bucket["topic_ids"]:
        return

    bucket["topic_ids"].add(topic_id)
    bucket["topic_count"] += 1
    bucket["reply_total"] += detail_row["weekly_reply_count"]
    bucket["vote_total"] += int(detail_row.get("vote_count", 0) or 0)
    if detail_row["title"] not in bucket["sample_titles"] and len(bucket["sample_titles"]) < 3:
        bucket["sample_titles"].append(detail_row["title"])


def rank_bucket_rows(buckets: dict[str, dict[str, Any]], limit: int, *, mode: str) -> list[dict[str, Any]]:
    if mode == "summary":
        ordered = sorted(
            buckets.values(),
            key=lambda item: (-item["topic_count"], -item["mention_count"], -item["reply_total"], item["label"]),
        )
    else:
        ordered = sorted(
            buckets.values(),
            key=lambda item: (-item["score"], -item["topic_count"], -item["reply_total"], item["label"]),
        )

    rows: list[dict[str, Any]] = []
    for rank, item in enumerate(ordered[:limit], start=1):
        rows.append(
            {
                "rank": rank,
                "label": item["label"],
                "topic_count": item["topic_count"],
                "reply_total": item["reply_total"],
                "vote_total": item["vote_total"],
                "mention_count": item["mention_count"],
                "score": item["score"],
                "representative_titles": "\n".join(item["sample_titles"]) or "-",
            }
        )
    return rows


def build_empty_ai_result_template(packet: dict[str, Any]) -> dict[str, Any]:
    return {
        "topic_id": packet["topic_id"],
        "title": packet["title"],
        "topic_url": packet["topic_url"],
        "ai_topic_category": "",
        "ai_is_negative_feedback": "",
        "ai_is_product_suggestion": "",
        "ai_core_summary": "",
        "notes": "",
    }


def build_seed_ai_result(packet: dict[str, Any]) -> dict[str, Any]:
    first_post = ""
    post_blocks = packet.get("post_blocks") or []
    if post_blocks:
        first_post = strip_html(post_blocks[0].get("content") or "")
    return {
        "topic_id": packet["topic_id"],
        "title": packet["title"],
        "topic_url": packet["topic_url"],
        "ai_topic_category": "",
        "ai_is_negative_feedback": "",
        "ai_is_product_suggestion": "",
        "ai_core_summary": clip_text(first_post, 80) if first_post else "",
        "notes": "seed-is-placeholder-needs-agent-judgement",
    }


def load_ai_results(ai_results_path: str | None) -> dict[int, dict[str, Any]]:
    if not ai_results_path:
        return {}
    payload = json.loads(Path(ai_results_path).read_text(encoding="utf-8"))
    if isinstance(payload, dict) and "results" in payload:
        records = payload["results"]
    else:
        records = payload
    mapping: dict[int, dict[str, Any]] = {}
    for item in records:
        topic_id = int(item["topic_id"])
        mapping[topic_id] = item
    return mapping


def update_ai_bucket(
    buckets: dict[str, dict[str, Any]],
    label: str,
    detail_row: dict[str, Any],
    *,
    score: int,
) -> None:
    bucket = buckets.setdefault(
        label,
        {
            "label": label,
            "topic_ids": set(),
            "topic_count": 0,
            "reply_total": 0,
            "vote_total": 0,
            "mention_count": 0,
            "score": 0,
            "sample_titles": [],
        },
    )
    bucket["mention_count"] += 1
    bucket["score"] += score
    if detail_row["topic_id"] in bucket["topic_ids"]:
        return
    bucket["topic_ids"].add(detail_row["topic_id"])
    bucket["topic_count"] += 1
    bucket["reply_total"] += detail_row["weekly_reply_count"]
    bucket["vote_total"] += int(detail_row.get("vote_count", 0) or 0)
    if detail_row["title"] not in bucket["sample_titles"] and len(bucket["sample_titles"]) < 3:
        bucket["sample_titles"].append(detail_row["title"])


def split_labels(value: str) -> list[str]:
    if not value or value == "-":
        return []
    return [part.strip() for part in re.split(r"[、,;/\n]+", value) if part.strip()]


def apply_ai_results_to_report(report: dict[str, Any], ai_results: dict[int, dict[str, Any]], top_n: int) -> dict[str, Any]:
    if not ai_results:
        report["ai_results_applied"] = False
        report["ai_results_count"] = 0
        return report

    ai_hot_buckets: dict[str, dict[str, Any]] = {}
    ai_negative_buckets: dict[str, dict[str, Any]] = {}
    ai_suggestion_buckets: dict[str, dict[str, Any]] = {}
    ai_results_count = 0
    hot_topic_exclusions = {"使用咨询与求助", "使用方式与工作流讨论"}

    for row in report["detail_rows"]:
        item = ai_results.get(int(row["topic_id"]))
        if not item:
            continue
        ai_results_count += 1
        row["ai_topic_category"] = item.get("ai_topic_category", "") or ""
        row["ai_is_negative_feedback"] = item.get("ai_is_negative_feedback", "") or ""
        row["ai_is_product_suggestion"] = item.get("ai_is_product_suggestion", "") or ""
        row["ai_core_summary"] = item.get("ai_core_summary", "") or ""

        score = max(1, int(row.get("weekly_post_count", 0))) + int(row.get("weekly_reply_count", 0)) + min(int(row.get("vote_count", 0) or 0), 3)
        for label in split_labels(row["ai_topic_category"]):
            if label not in hot_topic_exclusions:
                update_ai_bucket(ai_hot_buckets, label, row, score=score)
        if row["ai_is_negative_feedback"] == "是":
            labels = split_labels(row["ai_topic_category"]) or ["其他负向反馈"]
            for label in labels:
                update_ai_bucket(ai_negative_buckets, label, row, score=score)
        if row["ai_is_product_suggestion"] == "是":
            labels = split_labels(row["ai_topic_category"]) or ["其他产品建议"]
            for label in labels:
                update_ai_bucket(ai_suggestion_buckets, label, row, score=score)

    report["discussion_summaries"] = rank_bucket_rows(ai_hot_buckets, top_n, mode="summary")
    report["hot_topics"] = rank_bucket_rows(ai_hot_buckets, top_n, mode="hot")
    report["negative_feedback"] = rank_bucket_rows(ai_negative_buckets, top_n, mode="summary")
    report["product_suggestions"] = rank_bucket_rows(ai_suggestion_buckets, top_n, mode="summary")
    report["ai_results_applied"] = True
    report["ai_results_count"] = ai_results_count
    return report


def has_significant_hotspot(rows: list[dict[str, Any]]) -> bool:
    return any(row["topic_count"] >= 2 or row["reply_total"] >= 4 or row["vote_total"] >= 2 for row in rows)


def aggregate_report(
    base_url: str,
    start_local: datetime,
    end_local: datetime,
    max_pages: int,
    top_n: int,
) -> dict[str, Any]:
    start_utc = start_local.astimezone(timezone.utc)
    end_utc = end_local.astimezone(timezone.utc)
    limit = max(5, min(top_n, 5))
    created_topics = collect_created_topics(base_url, INTERACTIVE_CATEGORY_ID, start_utc, end_utc, max_pages)
    summary_buckets: dict[str, dict[str, Any]] = {}
    hot_buckets: dict[str, dict[str, Any]] = {}
    negative_buckets: dict[str, dict[str, Any]] = {}
    suggestion_buckets: dict[str, dict[str, Any]] = {}
    detail_rows: list[dict[str, Any]] = []
    analyzed_topics = 0
    activity_posts = 0
    analysis_packets: list[dict[str, Any]] = []

    for topic_stub in created_topics:
        if topic_stub.get("archetype") != "regular":
            continue
        if topic_stub.get("visible") is False:
            continue

        topic = fetch_topic_posts(topic_stub["id"], base_url)
        window_posts = topic_posts_within_window(topic, start_utc, end_utc)
        if not window_posts:
            continue

        analyzed_topics += 1
        title = topic.get("title") or topic_stub.get("title") or ""
        origin_text = strip_html((topic.get("all_posts") or [{}])[0].get("cooked") or "")
        reply_texts = [strip_html(post.get("cooked") or "") for post in window_posts if post.get("post_number", 0) > 1]
        weekly_post_count = len(window_posts)
        weekly_reply_count = sum(1 for post in window_posts if post.get("post_number", 0) > 1)
        vote_count = int(topic.get("vote_count") or 0)
        activity_posts += weekly_post_count
        weight = max(1, weekly_post_count) + weekly_reply_count + min(vote_count, 3)
        topic_url = f"{base_url.rstrip('/')}/t/-/{topic['id']}"
        created_at_local = parse_iso_datetime(topic["created_at"]).astimezone(ZoneInfo(DEFAULT_TIMEZONE)).strftime("%Y-%m-%d %H:%M:%S")
        post_blocks, full_thread_text = build_post_blocks(topic["all_posts"], DEFAULT_TIMEZONE)
        agent_analysis_input = build_agent_analysis_input(
            topic_id=topic["id"],
            title=title,
            topic_url=topic_url,
            created_at=created_at_local,
            views=int(topic.get("views") or 0),
            reply_count=int(topic.get("reply_count") or 0),
            weekly_post_count=weekly_post_count,
            weekly_reply_count=weekly_reply_count,
            thread_text=full_thread_text,
        )

        summary_terms = extract_pattern_terms(title, [origin_text, *reply_texts])
        hot_terms = extract_pattern_terms(title, [])
        keyword_terms = extract_keywords(title, [origin_text, *reply_texts])
        negative_terms = extract_negative_signals(" ".join(part for part in [title, origin_text, *reply_texts] if part))
        suggestion_terms = extract_product_suggestions(title)

        detail_row = {
            "topic_id": topic["id"],
            "title": title,
            "topic_url": topic_url,
            "created_at": created_at_local,
            "views": int(topic.get("views") or 0),
            "reply_count": int(topic.get("reply_count") or 0),
            "weekly_post_count": weekly_post_count,
            "weekly_reply_count": weekly_reply_count,
            "vote_count": vote_count,
            "summary_topics": "、".join(summary_terms) or "-",
            "keywords": "、".join(keyword_terms) or "-",
            "hot_topics": "、".join(hot_terms) or "-",
            "negative_feedback": "、".join(negative_terms) or "-",
            "product_suggestions": "、".join(suggestion_terms) or "-",
            "sample_excerpt": (reply_texts[0] if reply_texts else origin_text)[:180] or "-",
            "first_post_text": clip_text(origin_text, 600),
            "full_thread_text": clip_text(full_thread_text, 4000),
            "ai_topic_category": "",
            "ai_is_negative_feedback": "",
            "ai_is_product_suggestion": "",
            "ai_core_summary": "",
            "agent_analysis_input": agent_analysis_input,
        }
        detail_rows.append(detail_row)
        analysis_packets.append(
            {
                "topic_id": topic["id"],
                "title": title,
                "topic_url": topic_url,
                "created_at": created_at_local,
                "views": int(topic.get("views") or 0),
                "reply_count": int(topic.get("reply_count") or 0),
                "weekly_post_count": weekly_post_count,
                "weekly_reply_count": weekly_reply_count,
                "post_blocks": post_blocks,
                "full_thread_text": full_thread_text,
                "agent_analysis_input": agent_analysis_input,
                "rule_hints": {
                    "summary_topics": summary_terms,
                    "keywords": keyword_terms,
                    "hot_topics": hot_terms,
                    "negative_feedback": negative_terms,
                    "product_suggestions": suggestion_terms,
                },
            }
        )

        for label in summary_terms[:5]:
            update_bucket(summary_buckets, label, detail_row, score=weight, mention_weight=max(1, weekly_post_count))
        for label in hot_terms[:5]:
            update_bucket(hot_buckets, label, detail_row, score=weight, mention_weight=max(1, weekly_reply_count + 1))
        for label in negative_terms:
            update_bucket(negative_buckets, label, detail_row, score=weight, mention_weight=max(1, weekly_reply_count + 1))
        for label in suggestion_terms:
            update_bucket(suggestion_buckets, label, detail_row, score=weight + vote_count, mention_weight=max(1, weekly_post_count))

    rule_summary_rows = rank_bucket_rows(summary_buckets, limit, mode="summary")
    rule_hot_rows = rank_bucket_rows(hot_buckets, limit, mode="hot")
    rule_negative_rows = rank_bucket_rows(negative_buckets, limit, mode="summary")
    rule_suggestion_rows = rank_bucket_rows(suggestion_buckets, limit, mode="summary")
    fragmented = not has_significant_hotspot(rule_hot_rows)

    return {
        "forum_base_url": base_url.rstrip("/"),
        "strategy": "interaction-created-full-scan",
        "strategy_label": "互动交流按创建时间全量翻页扫描（规则候选版）",
        "target_categories": [{"id": INTERACTIVE_CATEGORY_ID, "name": INTERACTIVE_CATEGORY_NAME}],
        "summary": {
            "candidate_topics": len(created_topics),
            "analyzed_topics": analyzed_topics,
            "activity_posts": activity_posts,
            "fragmented": fragmented,
            "summary_cluster_count": len(summary_buckets),
            "hot_cluster_count": len(hot_buckets),
            "negative_cluster_count": len(negative_buckets),
            "suggestion_cluster_count": len(suggestion_buckets),
        },
        "discussion_summaries": [],
        "hot_topics": [],
        "negative_feedback": [],
        "product_suggestions": [],
        "fragmented_message": "待 Agent 判别" if not fragmented else "待 Agent 判别",
        "rule_candidates": {
            "discussion_summaries": rule_summary_rows,
            "hot_topics": rule_hot_rows,
            "negative_feedback": rule_negative_rows,
            "product_suggestions": rule_suggestion_rows,
        },
        "detail_rows": detail_rows,
        "analysis_packets": analysis_packets,
    }


def style_range(
    ws,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    *,
    fill=None,
    font=None,
    border=None,
    alignment=None,
) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def write_section(
    ws,
    start_row: int,
    title: str,
    rows: list[dict[str, Any]],
    headers: list[str],
    value_keys: list[str],
    *,
    navy: str,
    blue: str,
    white: str,
    text_dark: str,
    font_name: str,
    card_border: Border,
    header_border: Border,
    top1_fill,
    top2_fill,
    top3_fill,
    zebra_fill,
) -> int:
    section_fill = PatternFill("solid", fgColor=navy)
    header_fill = PatternFill("solid", fgColor=blue)

    ws.cell(start_row, 1, title)
    style_range(
        ws,
        start_row,
        1,
        start_row,
        len(headers),
        fill=section_fill,
        font=Font(name=font_name, size=13, bold=True, color=white),
        alignment=Alignment(horizontal="centerContinuous", vertical="center"),
    )

    header_row = start_row + 1
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.fill = header_fill
        cell.font = Font(name=font_name, size=10, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    if not rows:
        rows = [
            {
                "rank": "-",
                "label": "待 Agent 判别",
                "topic_count": 0,
                "reply_total": 0,
                "mention_count": 0,
                "vote_total": 0,
                "representative_titles": "-",
            }
        ]

    row_index = header_row + 1
    for row in rows:
        fill = zebra_fill
        if row["rank"] == 1:
            fill = top1_fill
        elif row["rank"] == 2:
            fill = top2_fill
        elif row["rank"] == 3:
            fill = top3_fill

        values = [row.get(key, "-") for key in value_keys]
        max_lines = 1
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col, value)
            cell.fill = fill
            cell.border = card_border
            cell.font = Font(name=font_name, size=10, color=text_dark, bold=bool(row.get("rank") in {1, 2, 3} and col == 2))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            max_lines = max(max_lines, str(value).count("\n") + 1)
        ws.row_dimensions[row_index].height = max(24, max_lines * 17)
        row_index += 1

    return row_index + 1


def write_summary_workbook(report: dict[str, Any], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"
    ai_applied = bool(report.get("ai_results_applied"))

    navy = "1F3A5F"
    blue = "2F75B5"
    white = "FFFFFF"
    text_dark = "1F1F1F"
    light_blue = "EAF3FF"
    card_fill = PatternFill("solid", fgColor=light_blue)
    top1_fill = PatternFill("solid", fgColor="FFF4CC")
    top2_fill = PatternFill("solid", fgColor="EDEDED")
    top3_fill = PatternFill("solid", fgColor="FCE4D6")
    zebra_fill = PatternFill("solid", fgColor="F5F7FA")
    font_name = "Microsoft YaHei"
    thin_gray = Side(style="thin", color="D9E2F3")
    medium_blue = Side(style="medium", color=blue)
    card_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_border = Border(left=medium_blue, right=medium_blue, top=medium_blue, bottom=medium_blue)

    widths = {"A": 8, "B": 28, "C": 14, "D": 14, "E": 14, "F": 46}
    for key, width in widths.items():
        ws.column_dimensions[key].width = width

    ws["A1"] = "互动交流热门讨论报告"
    style_range(
        ws,
        1,
        1,
        1,
        6,
        fill=PatternFill("solid", fgColor=navy),
        font=Font(name=font_name, size=18, bold=True, color=white),
        alignment=Alignment(horizontal="centerContinuous", vertical="center"),
    )
    ws["A2"] = (
        f"统计范围（{report['timezone']}）：{report['start_local'].strftime('%Y-%m-%d %H:%M:%S')} "
        f"至 {report['end_local'].strftime('%Y-%m-%d %H:%M:%S')}"
    )
    style_range(
        ws,
        2,
        1,
        2,
        6,
        fill=PatternFill("solid", fgColor="DCE6F1"),
        font=Font(name=font_name, size=11, color=text_dark),
        alignment=Alignment(horizontal="centerContinuous", vertical="center"),
    )

    ws["A3"] = f"统计板块：{INTERACTIVE_CATEGORY_NAME}"
    ws["B3"] = f"候选话题数：{report['summary']['candidate_topics']}"
    ws["C3"] = f"有效话题数：{report['summary']['analyzed_topics']}"
    ws["D3"] = f"窗口内帖子数：{report['summary']['activity_posts']}"
    ws["E3"] = f"规则候选聚类数：{report['summary']['summary_cluster_count']}"
    ws["F3"] = "明细请查看第二个页签“明细”，汇总结果仅在 Agent 回填后生成"
    style_range(
        ws,
        3,
        1,
        3,
        6,
        fill=card_fill,
        font=Font(name=font_name, size=10, bold=True, color=navy),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    ws["A4"] = "当前报表默认不输出伪 AI 汇总，需基于逐帖 Agent 判别结果生成 TOP 榜单"
    ws["B4"] = "已应用 Agent 判别结果" if ai_applied else "待 Agent 判别"
    ws["C4"] = f"热门主题聚类数：{report['summary']['hot_cluster_count']}"
    ws["D4"] = f"负向反馈聚类数：{report['summary']['negative_cluster_count']}"
    ws["E4"] = f"产品建议聚类数：{report['summary']['suggestion_cluster_count']}"
    ws["F4"] = f"数据源：{report['forum_base_url']}/c/11-category/11"
    style_range(
        ws,
        4,
        1,
        4,
        6,
        fill=PatternFill("solid", fgColor="F7FBFF"),
        font=Font(name=font_name, size=10, color=text_dark),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )

    section_rows = report["discussion_summaries"] if ai_applied else []
    next_row = write_section(
        ws,
        6,
        "1. 热门话题主题 TOP5",
        section_rows,
        ["排名", "讨论内容", "命中帖子数", "提及热度", "本周回帖数", "代表帖子"],
        ["rank", "label", "topic_count", "mention_count", "reply_total", "representative_titles"],
        navy=navy,
        blue=blue,
        white=white,
        text_dark=text_dark,
        font_name=font_name,
        card_border=card_border,
        header_border=header_border,
        top1_fill=top1_fill,
        top2_fill=top2_fill,
        top3_fill=top3_fill,
        zebra_fill=zebra_fill,
    )
    section_rows = report["negative_feedback"] if ai_applied else []
    next_row = write_section(
        ws,
        next_row,
        "2. 高频负向反馈 TOP5",
        section_rows,
        ["排名", "负向反馈", "命中帖子数", "提及热度", "本周回帖数", "代表帖子"],
        ["rank", "label", "topic_count", "mention_count", "reply_total", "representative_titles"],
        navy=navy,
        blue=blue,
        white=white,
        text_dark=text_dark,
        font_name=font_name,
        card_border=card_border,
        header_border=header_border,
        top1_fill=top1_fill,
        top2_fill=top2_fill,
        top3_fill=top3_fill,
        zebra_fill=zebra_fill,
    )
    section_rows = report["product_suggestions"] if ai_applied else []
    next_row = write_section(
        ws,
        next_row,
        "3. 产品建议 TOP5",
        section_rows,
        ["排名", "产品建议", "命中帖子数", "提及热度", "本周回帖数", "代表帖子"],
        ["rank", "label", "topic_count", "mention_count", "reply_total", "representative_titles"],
        navy=navy,
        blue=blue,
        white=white,
        text_dark=text_dark,
        font_name=font_name,
        card_border=card_border,
        header_border=header_border,
        top1_fill=top1_fill,
        top2_fill=top2_fill,
        top3_fill=top3_fill,
        zebra_fill=zebra_fill,
    )
    detail_ws = wb.create_sheet("明细")
    detail_ws.sheet_view.showGridLines = False
    detail_ws.freeze_panes = "A2"
    detail_headers = [
        "topic_id",
        "标题",
        "链接",
        "创建时间",
        "浏览量",
        "总回复数",
        "窗口内帖子数",
        "窗口内回帖数",
        "讨论内容分类",
        "关键词提取",
        "热门主题分类",
        "负向反馈分类",
        "产品建议分类",
        "AI话题分类",
        "AI是否负向反馈",
        "AI是否产品建议",
        "AI核心内容",
        "样本文本",
    ]
    detail_keys = [
        "topic_id",
        "title",
        "topic_url",
        "created_at",
        "views",
        "reply_count",
        "weekly_post_count",
        "weekly_reply_count",
        "summary_topics",
        "keywords",
        "hot_topics",
        "negative_feedback",
        "product_suggestions",
        "ai_topic_category",
        "ai_is_negative_feedback",
        "ai_is_product_suggestion",
        "ai_core_summary",
        "sample_excerpt",
    ]
    widths = [10, 36, 28, 20, 10, 10, 12, 12, 24, 24, 24, 24, 24, 18, 18, 18, 28, 44]
    for idx, width in enumerate(widths, start=1):
        if idx <= 26:
            column_name = chr(64 + idx)
        else:
            column_name = "A" + chr(64 + idx - 26)
        detail_ws.column_dimensions[column_name].width = width

    for col, header in enumerate(detail_headers, start=1):
        cell = detail_ws.cell(1, col, header)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(name=font_name, size=10, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    for row_index, row in enumerate(report["detail_rows"], start=2):
        fill = zebra_fill if row_index % 2 == 0 else card_fill
        max_lines = 1
        for col, key in enumerate(detail_keys, start=1):
            value = row.get(key, "-")
            cell = detail_ws.cell(row_index, col, value)
            cell.fill = fill
            cell.border = card_border
            cell.font = Font(name=font_name, size=10, color=text_dark)
            cell.alignment = Alignment(horizontal="left" if col in {2, 3, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18} else "center", vertical="center", wrap_text=True)
            max_lines = max(max_lines, str(value).count("\n") + 1)
        detail_ws.row_dimensions[row_index].height = max(22, max_lines * 17)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def serialize_report(report: dict[str, Any]) -> dict[str, Any]:
    return {
        "forum_base_url": report["forum_base_url"],
        "timezone": report["timezone"],
        "time_window": {
            "start": report["start_local"].isoformat(),
            "end": report["end_local"].isoformat(),
            "label": report["time_label"],
        },
        "strategy": report["strategy"],
        "strategy_label": report["strategy_label"],
        "target_categories": report["target_categories"],
        "summary": report["summary"],
        "discussion_summaries": report["discussion_summaries"],
        "hot_topics": report["hot_topics"],
        "negative_feedback": report["negative_feedback"],
        "product_suggestions": report["product_suggestions"],
        "rule_candidates": report.get("rule_candidates", {}),
        "detail_rows": report["detail_rows"],
        "analysis_packets": report["analysis_packets"],
        "ai_workflow": {
            "analysis_source": "current-agent",
            "topic_input_scope": "title-and-full-thread",
            "ai_results_applied": report.get("ai_results_applied", False),
            "ai_results_count": report.get("ai_results_count", 0),
            "expected_fields": [
                "ai_topic_category",
                "ai_is_negative_feedback",
                "ai_is_product_suggestion",
                "ai_core_summary",
            ],
        },
        "fragmented_message": report["fragmented_message"],
        "excel_path": report["excel_path"],
    }


def default_output_path(output_arg: str | None, start_local: datetime, end_local: datetime) -> Path:
    if output_arg:
        return Path(output_arg)

    output_dir = Path.cwd() / DEFAULT_OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"TRAE社区_互动交流热门话题_{start_local.strftime('%Y%m%d')}-{end_local.strftime('%Y%m%d')}.xlsx"
    return output_dir / filename


def reserve_writable_output_path(preferred_path: Path) -> Path:
    if not preferred_path.exists():
        return preferred_path

    for index in range(1, 100):
        candidate = preferred_path.with_name(f"{preferred_path.stem}-{index}{preferred_path.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Unable to reserve output path for {preferred_path}")


def main() -> None:
    args = parse_args()
    start_local, end_local, label = compute_window(args)
    output_path = reserve_writable_output_path(default_output_path(args.output, start_local, end_local))
    report = aggregate_report(
        args.forum_base_url.rstrip("/"),
        start_local,
        end_local,
        args.max_pages,
        args.top_n,
    )

    report["timezone"] = args.timezone
    report["time_label"] = label
    report["start_local"] = start_local
    report["end_local"] = end_local
    report["excel_path"] = str(output_path)
    report["ai_results_applied"] = False
    report["ai_results_count"] = 0
    ai_results = load_ai_results(args.ai_results)
    if ai_results:
        report = apply_ai_results_to_report(report, ai_results, args.top_n)
    write_summary_workbook(report, output_path)

    json_path = output_path.with_suffix(".json")
    json_path.write_text(json.dumps(serialize_report(report), ensure_ascii=False, indent=2), encoding="utf-8")
    if args.export_ai_template:
        template_path = output_path.with_name(output_path.stem + ".ai-template.json")
        review_guide_path = output_path.with_name(output_path.stem + ".ai-review.md")
        templates = [build_empty_ai_result_template(packet) for packet in report["analysis_packets"]]
        template_payload = {
            "meta": {
                "source_excel": str(output_path),
                "source_json": str(json_path),
                "topic_input_scope": "title-and-full-thread",
                "analysis_source": "current-agent",
                "topic_count": len(templates),
            },
            "results": templates,
        }
        review_guide = "\n".join(
            [
                "# Agent 判别说明",
                "",
                f"- 源 Excel: `{output_path}`",
                f"- 源 JSON: `{json_path}`",
                f"- 回填模板: `{template_path}`",
                "- 请逐条阅读 `analysis_packets` 中的 `agent_analysis_input`，基于标题和全帖内容做语义判断。",
                "- 不要依据标签，不要只做关键词匹配。",
                "- 每条仅回填以下字段：",
                "  - `ai_topic_category`: 话题核心主题，建议 1 个主类，必要时可写 2 个。",
                "  - `ai_is_negative_feedback`: 仅填写 `是` 或 `否`。",
                "  - `ai_is_product_suggestion`: 仅填写 `是` 或 `否`。",
                "  - `ai_core_summary`: 用一句话概括帖子的核心表达。",
                "- 回填完成后，将结果保存为 `ai_results.json`，再通过 `--ai-results` 重新生成最终汇总。",
            ]
        )
        template_path.write_text(json.dumps(template_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        review_guide_path.write_text(review_guide, encoding="utf-8")


if __name__ == "__main__":
    main()
